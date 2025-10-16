import logging.handlers
import random
import sys
import time

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from settings import *
from src.logger import set_logger
from src.utils import read_xlsx, seconds_to_hms, fetch_page

GOOD_STATUS = 'OK'

PURPLE_FILL = PatternFill(start_color="9966CC", end_color="9966CC", fill_type="solid")

PINK_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")


def set_status_to_table(wb: Workbook, ws: Worksheet, row_num: int, color: str):
    try:
        logging.debug(f'Обновление цвета в xlsx "{color}"...')
        # Подбор заливки только для фиолетовой/розовой зон
        if color == 'фиолетовая':
            selected_fill = PURPLE_FILL
        elif color == 'розовая':
            selected_fill = PINK_FILL
        else:
            selected_fill = None
        max_column = ws.max_column
        if selected_fill is not None:
            for col in range(1, max_column + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = selected_fill

        # Всегда записываем текст статуса в целевую колонку
        ws.cell(row=row_num, column=COLUMN_COLOR).value = color
        wb.save(FN_XLSX)
        return True
    except Exception as e:
        logging.error(f'Ошибка обновления цвета в таблице: {e}')
        return False


def check_coordinates(coordinates: str):
    # https://pvz-map-backend.wildberries.ru/api/v1/map/check-point?point.longitude=37.52181476187005&point.latitude=56.24526307172684&area=70&tariff_type=BasicTariff
    lat = coordinates.split('lat = ')[-1].split(',')[0]
    lng = coordinates.split('lng = ')[-1]

    # фиолетовая для тестов
    # lat = '56.24526307172684'
    # lng = '37.52181476187005'

    # розовая для тестов
    # lat = '56.21041715804586'
    # lng = '37.60666666091899'

    url = f'https://pvz-map-backend.wildberries.ru/api/v1/map/check-point?point.longitude={lng}&point.latitude={lat}&area={AREA}'
    response = fetch_page(url)
    if not response:
        return "ERROR"
    try:
        description = response.get("description", '')
        if description == 'Точка попадает в красную зону':
            return 'красная'
        if description == 'Точка попадает в здание с действующим или открывающимся ПВЗ':
            return 'с действующим или открывающимся ПВЗ'

        # Проверяем зоны через новую структуру API
        point_info = response.get("point_info", {})
        zone_info = point_info.get("zone_info", {})
        general_info = point_info.get("general_info", {})
        
        # Проверяем приоритетные зоны
        priority_zone_info = general_info.get("priority_zone_info", {})
        if priority_zone_info:
            text_code = priority_zone_info.get("text_code")
            if text_code == 'new_build_zones':
                return 'розовая'  # Розовая зона по населению
            elif text_code == 'load_predict':
                return 'фиолетовая'  # Фиолетовая зона (если есть)
        
        # Проверяем обычные зоны
        zone_text_code = zone_info.get("text_code")
        if zone_text_code == 'green_zone':
            return 'зеленая'  # Зеленая зона
    except Exception as e:
        logging.error(f'Ошибка разбора ответа от сервиса: {e}\nОтвет был: {response}')
    return 'N/A'


def main():
    logging.info('Начало работы')

    wb = read_xlsx(FN_XLSX, is_data_only=False)
    if not wb:
        sys.exit()
    ws = wb.worksheets[0]
    row_count = ws.max_row
    is_need_to_pause = False
    start_row = 2
    for row_num in range(start_row, row_count + 1):
        if is_need_to_pause:
            time.sleep(random.uniform(*PAUSE_REQUEST))

        coordinates = ws.cell(row_num, COLUMN_COORDINATES).value
        if not coordinates:
            continue

        prev_color = ws.cell(row_num, 2).value
        if prev_color and prev_color in ['фиолетовая', 'розовая']:
            logging.info(f'Координаты "{coordinates}" уже были ранее обработан. Пропускаем...')
            continue
        is_need_to_pause = True
        logging.info(f'{"-" * 30}> [{row_num}/{row_count}] Обрабатываются координаты "{coordinates}"')
        color = check_coordinates(coordinates)
        logging.info(f'Результат = {color}')
        # Нормализация статусов для записи
        if not color or color in ['ERROR', 'NO_POINT_INFO']:
            color_to_write = 'N/A'
        else:
            color_to_write = color
        set_status_to_table(wb=wb, ws=ws, row_num=row_num, color=color_to_write)


if __name__ == '__main__':
    start_time = time.perf_counter()
    set_logger()
    main()
    logging.info(f'Работа завершена.\n'
                 f'Затраченное время: {seconds_to_hms(round(time.perf_counter() - start_time, 2))}')
